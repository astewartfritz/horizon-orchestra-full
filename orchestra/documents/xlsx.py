"""Horizon Orchestra — Excel Generation.

Create Excel workbooks from raw data, DataFrames, or structured inputs.
Supports multiple sheets, charts, formatting, and merging.  Uses
``openpyxl`` with a try/except guard.

Usage::

    from orchestra.documents.xlsx import XLSXGenerator

    gen = XLSXGenerator()
    data = [["Name", "Score"], ["Alice", 95], ["Bob", 87]]
    xlsx_bytes = gen.from_data(data, headers=True)
"""

from __future__ import annotations

import io
import logging
import os
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional, Sequence, Union

__all__ = [
    "XLSXGenerator",
]

log = logging.getLogger("orchestra.documents.xlsx")

_WORKSPACE = Path(os.environ.get("ORCHESTRA_WORKSPACE", "/tmp/orchestra_docs"))

# Optional dependency: openpyxl
try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        numbers as xl_numbers,
    )
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, AreaChart, ScatterChart, Reference,
    )
    from openpyxl.utils import get_column_letter
    _HAS_OPENPYXL = True
except ImportError:
    openpyxl = None  # type: ignore[assignment]
    _HAS_OPENPYXL = False


# ---------------------------------------------------------------------------
# Style presets
# ---------------------------------------------------------------------------

@dataclass
class StylePreset:
    """Style configuration for Excel formatting."""

    header_font: str = "Calibri"
    header_size: int = 12
    header_bold: bool = True
    header_bg: str = "1E3A5F"
    header_fg: str = "FFFFFF"
    body_font: str = "Calibri"
    body_size: int = 11
    alt_row_bg: str = "F5F7FA"
    border_color: str = "D0D5DD"
    number_format: str = "#,##0"
    currency_format: str = "$#,##0.00"
    percent_format: str = "0.0%"
    date_format: str = "YYYY-MM-DD"


# ---------------------------------------------------------------------------
# XLSXGenerator
# ---------------------------------------------------------------------------

class XLSXGenerator:
    """Excel workbook generator using openpyxl.

    Parameters
    ----------
    workspace:
        Directory for saving output files.
    style:
        Default style preset.
    """

    def __init__(
        self,
        workspace: str | Path | None = None,
        style: StylePreset | None = None,
    ) -> None:
        if not _HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel generation: pip install openpyxl"
            )
        self.workspace = Path(workspace) if workspace else _WORKSPACE / "xlsx"
        self.workspace.mkdir(parents=True, exist_ok=True)
        self.style = style or StylePreset()

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _save_path(self) -> Path:
        return self.workspace / f"{uuid.uuid4().hex[:12]}.xlsx"

    def _apply_header_style(self, ws: Any, row: int = 1, ncols: int = 0) -> None:
        """Apply header styling to a row."""
        s = self.style
        header_font = Font(
            name=s.header_font,
            size=s.header_size,
            bold=s.header_bold,
            color=s.header_fg,
        )
        header_fill = PatternFill(
            start_color=s.header_bg,
            end_color=s.header_bg,
            fill_type="solid",
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            bottom=Side(style="thin", color=s.border_color),
        )

        for col in range(1, ncols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _apply_body_style(self, ws: Any, start_row: int, end_row: int, ncols: int) -> None:
        """Apply body styling with alternating row colors."""
        s = self.style
        body_font = Font(name=s.body_font, size=s.body_size)
        alt_fill = PatternFill(
            start_color=s.alt_row_bg,
            end_color=s.alt_row_bg,
            fill_type="solid",
        )
        thin_border = Border(
            bottom=Side(style="hair", color=s.border_color),
        )

        for row in range(start_row, end_row + 1):
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = body_font
                cell.border = thin_border
                if (row - start_row) % 2 == 1:
                    cell.fill = alt_fill

    def _auto_width(self, ws: Any) -> None:
        """Auto-fit column widths based on content."""
        for col in ws.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    if cell_len > max_length:
                        max_length = cell_len
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, 8), 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _wb_to_bytes(self, wb: Any) -> bytes:
        """Convert a workbook to bytes."""
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def from_data(
        self,
        data: Sequence[Sequence[Any]],
        *,
        headers: bool = True,
        sheet_name: str = "Sheet1",
        apply_style: bool = True,
    ) -> bytes:
        """Create an Excel workbook from a 2D data array.

        Parameters
        ----------
        data:
            2D array of values.  First row is headers if *headers=True*.
        headers:
            Whether the first row contains headers.
        sheet_name:
            Worksheet name.
        apply_style:
            Whether to apply styling.

        Returns
        -------
        bytes
            XLSX file content.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        for row in data:
            ws.append(list(row))

        if apply_style and data:
            ncols = max(len(row) for row in data) if data else 0
            if headers and ncols:
                self._apply_header_style(ws, row=1, ncols=ncols)
                if len(data) > 1:
                    self._apply_body_style(ws, start_row=2, end_row=len(data), ncols=ncols)
            self._auto_width(ws)
            # Freeze header row
            if headers:
                ws.freeze_panes = "A2"

        return self._wb_to_bytes(wb)

    def from_dataframe(
        self,
        df: Any,
        *,
        sheet_name: str = "Sheet1",
        index: bool = False,
        apply_style: bool = True,
    ) -> bytes:
        """Create an Excel workbook from a pandas DataFrame.

        Parameters
        ----------
        df:
            pandas DataFrame.
        sheet_name:
            Worksheet name.
        index:
            Whether to include the DataFrame index.
        apply_style:
            Whether to apply styling.

        Returns
        -------
        bytes
            XLSX file content.
        """
        # Convert DataFrame to list of lists
        headers = list(df.columns)
        if index:
            headers = [df.index.name or "Index"] + headers

        rows: list[list[Any]] = [headers]
        for idx, row in df.iterrows():
            row_data = list(row)
            if index:
                row_data = [idx] + row_data
            rows.append(row_data)

        return self.from_data(rows, headers=True, sheet_name=sheet_name, apply_style=apply_style)

    def add_chart(
        self,
        xlsx_data: bytes,
        chart_type: str = "bar",
        data_range: str = "",
        *,
        sheet_name: str | None = None,
        title: str = "",
        x_title: str = "",
        y_title: str = "",
        min_col: int = 1,
        max_col: int = 0,
        min_row: int = 1,
        max_row: int = 0,
        categories_col: int = 1,
    ) -> bytes:
        """Add a chart to an existing workbook.

        Parameters
        ----------
        xlsx_data:
            Existing XLSX content.
        chart_type:
            Chart type (``bar``, ``line``, ``pie``, ``area``, ``scatter``).
        data_range:
            Excel-style range (e.g., ``"A1:C10"``).  Ignored if min/max
            params are provided.
        sheet_name:
            Target worksheet name.
        title:
            Chart title.
        x_title:
            X-axis title.
        y_title:
            Y-axis title.
        min_col, max_col, min_row, max_row:
            Data range boundaries.
        categories_col:
            Column number containing category labels.

        Returns
        -------
        bytes
            Updated XLSX content.
        """
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_data))
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        # Auto-detect range if not specified
        if not max_col:
            max_col = ws.max_column or 2
        if not max_row:
            max_row = ws.max_row or 10

        chart_map = {
            "bar": BarChart,
            "line": LineChart,
            "pie": PieChart,
            "area": AreaChart,
            "scatter": ScatterChart,
        }
        ChartClass = chart_map.get(chart_type, BarChart)
        chart = ChartClass()

        if title:
            chart.title = title
        if x_title and hasattr(chart, "x_axis"):
            chart.x_axis.title = x_title
        if y_title and hasattr(chart, "y_axis"):
            chart.y_axis.title = y_title

        chart.style = 10
        chart.width = 18
        chart.height = 12

        # Create data reference
        data_ref = Reference(
            ws,
            min_col=categories_col + 1,
            min_row=min_row,
            max_col=max_col,
            max_row=max_row,
        )
        cats_ref = Reference(
            ws,
            min_col=categories_col,
            min_row=min_row + 1,
            max_row=max_row,
        )

        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)

        # Place chart below the data
        chart_cell = f"A{max_row + 2}"
        ws.add_chart(chart, chart_cell)

        return self._wb_to_bytes(wb)

    def add_formatting(
        self,
        xlsx_data: bytes,
        styles: dict[str, Any],
        *,
        sheet_name: str | None = None,
    ) -> bytes:
        """Apply formatting rules to a workbook.

        Parameters
        ----------
        xlsx_data:
            Existing XLSX content.
        styles:
            Formatting rules dict. Supported keys:

            - ``"number_columns"`` — list of column letters for number formatting
            - ``"currency_columns"`` — list of column letters for currency formatting
            - ``"percent_columns"`` — list of column letters for percentage formatting
            - ``"date_columns"`` — list of column letters for date formatting
            - ``"bold_rows"`` — list of row numbers to bold
            - ``"highlight_cells"`` — dict of cell references → fill colors
        sheet_name:
            Target worksheet.

        Returns
        -------
        bytes
            Formatted XLSX content.
        """
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_data))
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        # Number format columns
        for col_letter in styles.get("number_columns", []):
            for row in range(2, (ws.max_row or 1) + 1):
                ws[f"{col_letter}{row}"].number_format = self.style.number_format

        for col_letter in styles.get("currency_columns", []):
            for row in range(2, (ws.max_row or 1) + 1):
                ws[f"{col_letter}{row}"].number_format = self.style.currency_format

        for col_letter in styles.get("percent_columns", []):
            for row in range(2, (ws.max_row or 1) + 1):
                ws[f"{col_letter}{row}"].number_format = self.style.percent_format

        for col_letter in styles.get("date_columns", []):
            for row in range(2, (ws.max_row or 1) + 1):
                ws[f"{col_letter}{row}"].number_format = self.style.date_format

        # Bold rows
        for row_num in styles.get("bold_rows", []):
            for col in range(1, (ws.max_column or 1) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.font = Font(name=self.style.body_font, size=self.style.body_size, bold=True)

        # Highlight cells
        for cell_ref, color in styles.get("highlight_cells", {}).items():
            ws[cell_ref].fill = PatternFill(
                start_color=color.lstrip("#"),
                end_color=color.lstrip("#"),
                fill_type="solid",
            )

        return self._wb_to_bytes(wb)

    def merge_sheets(
        self,
        files: Sequence[str | Path | bytes],
        *,
        sheet_names: Sequence[str] | None = None,
    ) -> bytes:
        """Merge multiple Excel files into one workbook.

        Each source file's first sheet is added as a separate sheet
        in the output workbook.

        Parameters
        ----------
        files:
            List of XLSX file paths or bytes.
        sheet_names:
            Custom sheet names for each file.

        Returns
        -------
        bytes
            Merged XLSX content.
        """
        output_wb = openpyxl.Workbook()
        # Remove the default sheet
        default_ws = output_wb.active
        output_wb.remove(default_ws)

        for i, file_data in enumerate(files):
            if isinstance(file_data, bytes):
                src_wb = openpyxl.load_workbook(io.BytesIO(file_data))
            else:
                src_wb = openpyxl.load_workbook(str(file_data))

            src_ws = src_wb.active
            name = (sheet_names[i] if sheet_names and i < len(sheet_names)
                    else src_ws.title or f"Sheet{i + 1}")

            # Ensure unique name
            if name in output_wb.sheetnames:
                name = f"{name}_{i + 1}"

            dest_ws = output_wb.create_sheet(title=name)

            for row in src_ws.iter_rows(values_only=False):
                for cell in row:
                    dest_cell = dest_ws.cell(
                        row=cell.row, column=cell.column, value=cell.value,
                    )
                    if cell.has_style:
                        dest_cell.font = cell.font.copy()
                        dest_cell.fill = cell.fill.copy()
                        dest_cell.alignment = cell.alignment.copy()
                        dest_cell.border = cell.border.copy()
                        dest_cell.number_format = cell.number_format

            # Copy column widths
            for col_letter, dim in src_ws.column_dimensions.items():
                dest_ws.column_dimensions[col_letter].width = dim.width

            src_wb.close()

        return self._wb_to_bytes(output_wb)

    def save(self, xlsx_bytes: bytes, path: str | Path | None = None) -> Path:
        """Save XLSX bytes to a file.

        Parameters
        ----------
        xlsx_bytes:
            XLSX content.
        path:
            Output path.

        Returns
        -------
        Path
            Saved file path.
        """
        save_to = Path(path) if path else self._save_path()
        save_to.parent.mkdir(parents=True, exist_ok=True)
        save_to.write_bytes(xlsx_bytes)
        log.info("Saved XLSX → %s (%d bytes)", save_to, len(xlsx_bytes))
        return save_to
